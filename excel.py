import openpyxl
import platform

from openpyxl import load_workbook
import extra_functions
from openpyxl.utils.exceptions import IllegalCharacterError

excel_product_image_url = []
excel_product_names = []
work_sheet_index = 2
table_location = 'excel\\a.xlsx' if platform.system() == 'Windows' else 'excel/a.xlsx'

product_ids = []
product_codes = []


def get_track_list_table(product_id_column, product_code_column):
    global product_ids
    global product_codes

    product_ids = get_all_the_rows_from_column(product_id_column)
    product_codes = get_all_the_rows_from_column(product_code_column)
    print('table entries: ', len(product_ids))


def sanitise_product_names(string):
    """ Replacing all the bad characters that inhibit search"""
    replace_commas = string.replace(',', '') if string.find(',') != -1 else string
    replace_stars = replace_commas.replace('*', ' ') if replace_commas.find('*') != -1 else replace_commas
    replace_slashes = replace_stars.replace(r'/', ' ') if replace_stars.find(r'/') != -1 else replace_stars
    replace_dollar_signs = replace_slashes.replace(' $', '') if replace_slashes.find(
        ' $') != -1 else replace_slashes
    replace_plus_sign = replace_dollar_signs.replace('+', ' ') if replace_dollar_signs.find(
        '+') != -1 else replace_dollar_signs
    replace_dots = replace_plus_sign.replace('.', ' ') if replace_plus_sign.find('.') != -1 else replace_plus_sign
    replace_dashes = replace_dots.replace('-', ' ') if replace_dots.find('-') != -1 else replace_dots
    replace_small = replace_dashes.replace('<', ' ') if replace_dashes.find('<') != -1 else replace_dashes
    replace_big = replace_small.replace('>', ' ') if replace_small.find('>') != -1 else replace_small
    replace_percentage = replace_big.replace(r"%", ' ') if replace_big.find(r"%") != -1 else replace_big
    return replace_percentage


def write_description_to_excel(table_string, description_column):
    global work_sheet_index

    wb = openpyxl.load_workbook(table_location)
    ws = wb.active

    product_brand_key = extra_functions.value_key(description_column, work_sheet_index)
    ws[product_brand_key] = table_string

    wb.save(table_location)


def sanitise_string(string):
    if isinstance(string, str):
        sanitised_string = string.encode('unicode_escape').decode('utf-8')
        return sanitised_string
    else:
        return string


def read_image_urls(cell_letter):
    global excel_product_image_url
    get_work_sheet_index()
    excel_product_image_url = get_all_the_rows_from_column(cell_letter)


def read_product_names(cell_letter):
    global excel_product_names
    get_work_sheet_index()
    excel_product_names = get_all_the_rows_from_column(cell_letter)


def get_all_the_rows_from_column(cell_letter):
    wb = load_workbook(table_location)  # Work Book
    ws = wb.get_sheet_by_name('Sheet1')  # Work Sheet
    column = ws[cell_letter]  # Column
    column_list = [column[x].value for x in range(len(column))]
    return column_list


def get_all_the_rows_from_column2(cell_letter):
    wb = load_workbook(table_location_b)  # Work Book
    ws = wb.get_sheet_by_name('Sheet1')  # Work Sheet
    column = ws[cell_letter]  # Column
    column_list = [column[x].value for x in range(len(column))]
    return column_list


def get_work_sheet_index():
    global work_sheet_index
    wb = load_workbook(table_location)  # Work Book
    ws = wb.get_sheet_by_name('Sheet1')  # Work Sheet
    column = ws['A']  # Column
    work_sheet_index = len(column) + 1
    print('Worksheet index: ', work_sheet_index)
